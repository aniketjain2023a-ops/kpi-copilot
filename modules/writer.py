import pandas as pd
from pathlib import Path
from datetime import datetime
def generate_ai_recommendations(analysis_df):
    recommendations = []

    for _, row in analysis_df.iterrows():
        status = str(row.get("Status", ""))

        if status == "On Track":
            continue

        kpi_name = str(row.get("Parameter Name", ""))

        if status == "Critical":
            recommendation = (
                "Immediate management review and corrective action required."
            )
        else:
            recommendation = (
                "Monitor weekly and implement targeted improvement actions."
            )

        recommendations.append(f"{kpi_name}: {recommendation}")

    return recommendations

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def write_analysis_sheet(workbook_path, analysis_df):
    wb = load_workbook(workbook_path)
    print(f"[WRITER] Opening workbook: {workbook_path}")
    print(f"[WRITER] Incoming dataframe shape: {analysis_df.shape}")
    if 'Parameter Name' in analysis_df.columns and 'Achievement %' in analysis_df.columns:
        print('[WRITER] Achievement values being written:')
        print(
            analysis_df[['Parameter Name', 'Achievement %']]
                .to_string(index=False)
        )

    if "Calculations" in wb.sheetnames:
        wb.remove(wb["Calculations"])

    ws = wb.create_sheet("Calculations")

    ws.freeze_panes = "A4"

    ws["A1"] = "KPI COPILOT MANAGEMENT REPORT"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:F1")

    table_row_start = 4

    # Build trend column BEFORE writing table
    trend_count_declining = 0

    if "Achievement %" in analysis_df.columns and "Trend" not in analysis_df.columns:
        trends = []
        for _, row in analysis_df.iterrows():
            achievement = row["Achievement %"]

            if achievement >= 70:
                trend = "📈 Improving"
            elif achievement <= 60:
                trend = "📉 Declining"
                trend_count_declining += 1
            else:
                trend = "➖ Stable"

            trends.append(trend)

        analysis_df["Trend"] = trends

    # KPI table header
    for idx, col_name in enumerate(analysis_df.columns, start=1):
        cell = ws.cell(row=table_row_start, column=idx)
        cell.value = col_name
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    # KPI table data
    for r_idx, row in enumerate(
        dataframe_to_rows(analysis_df, index=False, header=False),
        start=table_row_start + 1,
    ):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Status colour formatting
    if "Status" in analysis_df.columns:
        status_col = list(analysis_df.columns).index("Status") + 1

        for row_num in range(table_row_start + 1, table_row_start + 1 + len(analysis_df)):
            cell = ws.cell(row=row_num, column=status_col)

            if cell.value == "On Track":
                cell.fill = PatternFill("solid", fgColor="92D050")
            elif cell.value == "Monitor":
                cell.fill = PatternFill("solid", fgColor="FFD966")
            elif cell.value == "Critical":
                cell.fill = PatternFill("solid", fgColor="FF6666")

    # Trend colour formatting
    if "Trend" in analysis_df.columns:
        trend_col = list(analysis_df.columns).index("Trend") + 1

        for row_num in range(table_row_start + 1, table_row_start + 1 + len(analysis_df)):
            cell = ws.cell(row=row_num, column=trend_col)

            if "Improving" in str(cell.value):
                cell.fill = PatternFill("solid", fgColor="92D050")
            elif "Stable" in str(cell.value):
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
            elif "Declining" in str(cell.value):
                cell.fill = PatternFill("solid", fgColor="FF6666")

    last_col = get_column_letter(len(analysis_df.columns))
    last_row = table_row_start + len(analysis_df)
    ws.auto_filter.ref = f"A{table_row_start}:{last_col}{last_row}"

    if "Trend" in analysis_df.columns:
        trend_count_declining = len(
            analysis_df[
                analysis_df["Trend"].astype(str).str.contains("Declining", na=False)
            ]
        )

    summary_start_row = table_row_start + len(analysis_df) + 4

    ws[f"A{summary_start_row}"] = "KPI PERFORMANCE SUMMARY"
    ws[f"A{summary_start_row}"].font = Font(
        size=14,
        bold=True,
        color="FFFFFF"
    )
    ws[f"A{summary_start_row}"].fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    ws[f"A{summary_start_row + 1}"] = (
        f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    )

    total_kpis = len(analysis_df)

    on_track = 0
    monitor = 0
    critical = 0

    if "Status" in analysis_df.columns:
        on_track = len(analysis_df[analysis_df["Status"] == "On Track"])
        monitor = len(analysis_df[analysis_df["Status"] == "Monitor"])
        critical = len(analysis_df[analysis_df["Status"] == "Critical"])

    ws[f"A{summary_start_row + 3}"] = f"Total KPIs: {total_kpis}"
    ws[f"A{summary_start_row + 4}"] = f"On Track: {on_track}"
    ws[f"A{summary_start_row + 5}"] = f"Monitor: {monitor}"
    ws[f"A{summary_start_row + 6}"] = f"Critical: {critical}"

    ws[f"A{summary_start_row + 4}"].fill = PatternFill("solid", fgColor="92D050")
    ws[f"A{summary_start_row + 5}"].fill = PatternFill("solid", fgColor="FFD966")
    ws[f"A{summary_start_row + 6}"].fill = PatternFill("solid", fgColor="FF6666")

    ws[f"A{summary_start_row + 7}"] = f"Declining KPIs: {trend_count_declining}"

    # Executive Narrative section
    risk_kpi = "N/A"
    if "Status" in analysis_df.columns:
        critical_rows = analysis_df[analysis_df["Status"] == "Critical"]
        if not critical_rows.empty:
            risk_kpi = critical_rows.iloc[0]["Parameter Name"]

    top_performer = "N/A"
    if "Achievement %" in analysis_df.columns:
        top_performer = analysis_df.loc[
            analysis_df["Achievement %"].idxmax()
        ]["Parameter Name"]

    narrative = (
        f"Overall KPI health is stable.\n\n"
        f"• {on_track} of {total_kpis} KPIs are on track.\n"
        f"• Highest risk KPI: {risk_kpi}.\n"
        f"• Top performer: {top_performer}.\n"
        f"• Declining KPIs: {trend_count_declining}.\n\n"
        f"Recommended Action:\nFocus on {risk_kpi} immediately."
    )

    ws[f"A{summary_start_row + 10}"] = "EXECUTIVE NARRATIVE"
    ws[f"A{summary_start_row + 10}"].font = Font(bold=True, color="FFFFFF")
    ws[f"A{summary_start_row + 10}"].fill = PatternFill("solid", fgColor="1F4E78")

    ws.merge_cells(
        start_row=summary_start_row + 11,
        start_column=1,
        end_row=summary_start_row + 14,
        end_column=4,
    )

    ws[f"A{summary_start_row + 11}"] = narrative
    ws[f"A{summary_start_row + 11}"].alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )
    ws.row_dimensions[summary_start_row + 11].height = 120

    # AI Insights
    ws[f"F{summary_start_row}"] = "AI INSIGHTS"
    ws.merge_cells(
        start_row=summary_start_row,
        start_column=6,
        end_row=summary_start_row,
        end_column=8,
    )
    ws[f"F{summary_start_row}"].font = Font(bold=True, color="FFFFFF")
    ws[f"F{summary_start_row}"].fill = PatternFill("solid", fgColor="1F4E78")

    ws[f"F{summary_start_row + 2}"] = f"Risk KPI: {risk_kpi}"
    ws[f"F{summary_start_row + 3}"] = f"Top Performer: {top_performer}"
    ws[f"F{summary_start_row + 4}"] = f"Declining KPIs: {trend_count_declining}"

    highest_effort = "N/A"
    if "Required Monthly Target" in analysis_df.columns:
        highest_effort = analysis_df.loc[
            analysis_df["Required Monthly Target"].idxmax()
        ]["Parameter Name"]

    lowest_kpi = "N/A"
    if "Achievement %" in analysis_df.columns:
        lowest_kpi = analysis_df.loc[
            analysis_df["Achievement %"].idxmin()
        ]["Parameter Name"]

    ws[f"F{summary_start_row + 5}"] = f"Highest Effort: {highest_effort}"
    ws[f"F{summary_start_row + 6}"] = f"Lowest Achievement: {lowest_kpi}"
    ws[f"F{summary_start_row + 7}"] = f"Priority Action: Focus on {risk_kpi}"

    ws[f"F{summary_start_row + 2}"].fill = PatternFill("solid", fgColor="FF6666")
    ws[f"F{summary_start_row + 3}"].fill = PatternFill("solid", fgColor="92D050")
    ws[f"F{summary_start_row + 5}"].fill = PatternFill("solid", fgColor="FFD966")
    ws[f"F{summary_start_row + 7}"].fill = PatternFill("solid", fgColor="1F4E78")
    ws[f"F{summary_start_row + 7}"].font = Font(color="FFFFFF", bold=True)

    # AI Recommendations Sheet
    print("[WRITER] Calculations sheet generated successfully")
    recommendations = generate_ai_recommendations(analysis_df)

    if "AI Recommendations" in wb.sheetnames:
        wb.remove(wb["AI Recommendations"])

    ai_ws = wb.create_sheet("AI Recommendations")

    ai_ws["A1"] = "AI GENERATED RECOMMENDATIONS"
    ai_ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ai_ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ai_ws.merge_cells("A1:F1")

    ai_ws["A3"] = "Generated"
    ai_ws["B3"] = datetime.now().strftime('%d %b %Y, %I:%M %p')

    ai_ws["A5"] = "Recommendations"
    ai_ws["A5"].font = Font(bold=True, color="FFFFFF")
    ai_ws["A5"].fill = PatternFill("solid", fgColor="1F4E78")

    rec_row = 7

    if recommendations:
        for rec in recommendations:
            ai_ws[f"A{rec_row}"] = rec
            ai_ws[f"A{rec_row}"].alignment = Alignment(wrap_text=True)
            rec_row += 2
    else:
        ai_ws["A7"] = "No recommendations generated."

    ai_ws.column_dimensions["A"].width = 120
    for row in range(7, rec_row + 1):
        ai_ws.row_dimensions[row].height = 40

    # Historical comparison section
    history_dir = Path("history")
    history_dir.mkdir(exist_ok=True)

    history_files = sorted(history_dir.glob("*.csv"))

    if history_files:
        try:
            previous_df = pd.read_csv(history_files[-1])

            ws[f"J{summary_start_row}"] = "HISTORICAL COMPARISON"
            ws[f"J{summary_start_row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"J{summary_start_row}"].fill = PatternFill("solid", fgColor="1F4E78")

            comparison_row = summary_start_row + 2

            if (
                "Parameter Name" in previous_df.columns
                and "Achievement %" in previous_df.columns
                and "Parameter Name" in analysis_df.columns
                and "Achievement %" in analysis_df.columns
            ):
                previous_map = previous_df.set_index("Parameter Name")["Achievement %"].to_dict()

                for _, row in analysis_df.iterrows():
                    name = row["Parameter Name"]
                    current = row["Achievement %"]

                    if name in previous_map:
                        old = previous_map[name]

                        if current > old:
                            movement = "↑ Improved"
                        elif current < old:
                            movement = "↓ Declined"
                        else:
                            movement = "→ No Change"

                        ws[f"J{comparison_row}"] = f"{name}: {movement}"
                        comparison_row += 1
        except Exception:
            pass

    # Light borders for report sections
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="D9D9D9")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

    snapshot_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".csv"
    analysis_df.to_csv(history_dir / snapshot_name, index=False)

    ai_ws.freeze_panes = "A6"

    # Force workbook save and verify path
    wb.save(workbook_path)
    wb.close()

    print(f"[WRITER] Workbook saved successfully: {workbook_path}")
    print(f"[WRITER] Analysis rows written: {len(analysis_df)}")
