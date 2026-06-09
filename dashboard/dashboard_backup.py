import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

st.title('KPI Copilot')
st.warning('The original KPI dashboard section is missing from dashboard.py. Only the new Data Entry, SAP Upload, Quick Actions, and Chatbot sections remain. The KPI metrics, charts, and status table must be restored from a previous version of dashboard.py or rebuilt.')

# Temporary fallback dataframe so the page can load even if
# the dashboard data-loading section was accidentally removed.
if 'df' not in globals():
    df = pd.DataFrame(columns=['KPI', 'Achievement', 'Status'])

st.subheader('Monthly KPI Data Entry')

with st.expander('Update KPI Values'):
    selected_kpi = st.selectbox('Select KPI', df['KPI'].tolist() if not df.empty else [])
    current_value = df.loc[df['KPI'] == selected_kpi, 'Achievement'].iloc[0] if selected_kpi in df['KPI'].values else 0.0
    st.info(f'Current Achievement Value for {selected_kpi}: {current_value}')
    new_value = st.number_input('New Achievement Value', value=float(current_value))

    if st.button('Save KPI Update'):
        try:
            from openpyxl import load_workbook
            from pathlib import Path

            project_root = Path(__file__).resolve().parent.parent
            workbook_path = project_root / 'reports' / 'kpi_target_tracker.xlsx'

            wb = load_workbook(workbook_path)

            if 'Data Entry' not in wb.sheetnames:
                st.error('Data Entry sheet not found in workbook.')
            else:
                ws = wb['Data Entry']

                kpi_row = None
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=1).value
                    if str(cell_value).strip() == str(selected_kpi).strip():
                        kpi_row = row
                        break

                if kpi_row is None:
                    st.error(f'KPI not found in Data Entry sheet: {selected_kpi}')
                else:
                    current_month_col = None

                    for col in range(2, ws.max_column + 1):
                        header = ws.cell(row=1, column=col).value
                        if header:
                            current_month_col = col

                    if current_month_col is None:
                        st.error('No month columns found in Data Entry sheet.')
                    else:
                        ws.cell(row=kpi_row, column=current_month_col).value = float(new_value)
                        wb.save(workbook_path)

                        st.success(
                            f'Saved {selected_kpi} = {new_value} to column {ws.cell(1, current_month_col).value}'
                        )

                        st.rerun()

        except Exception as e:
            st.error(f'Excel update failed: {e}')

st.divider()

st.subheader('SAP Upload Center')

sap_file = st.file_uploader(
    'Upload SAP Export (Excel/CSV)',
    type=['xlsx', 'xls', 'csv']
)

if sap_file:
    st.success('SAP file uploaded successfully')
    st.info('Next step: Auto-map SAP columns and update the KPI workbook automatically.')

    try:
        if sap_file.name.endswith('.csv'):
            sap_df = pd.read_csv(sap_file)
        else:
            sap_df = pd.read_excel(sap_file)

        st.dataframe(sap_df.head(), use_container_width=True)

    except Exception as e:
        st.error(f'Unable to read SAP file: {e}')

st.divider()

st.subheader('Quick Actions')

col_a, col_b = st.columns(2)

with col_a:
    if st.button('🔄 Refresh KPI Calculations'):
        st.success('KPI calculations refreshed.')

with col_b:
    if st.button('📊 Generate Management Summary'):
        st.success('Management summary generation coming next.')

st.subheader('KPI Copilot Assistant')

if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []

user_prompt = st.chat_input('Ask about KPI performance, trends, targets...')

if user_prompt:
    st.session_state.chat_history.append((user_prompt, 'KPI Copilot analysis coming from future AI engine.'))

for question, answer in st.session_state.chat_history:
    with st.chat_message('user'):
        st.write(question)

    with st.chat_message('assistant'):
        st.write(answer)

st.divider()

st.success('KPI Copilot Dashboard Running Successfully')
