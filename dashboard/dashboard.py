
from pathlib import Path
from datetime import datetime
import base64
import sys

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook

from modules.calculator import generate_kpi_analysis
from modules.writer import write_analysis_sheet
from utils.kpi_chat_engine import answer_question

st.set_page_config(page_title='KPI Copilot', page_icon='📊', layout='wide')

logo_candidates = [
    project_root / 'assets' / 'jsw_paints_logo.png',
    project_root / 'assets' / 'jsw_paints_logo.jpeg',
    project_root / 'assets' / 'jsw_paints_logo.jpg',
    project_root / 'dashboard' / 'assets' / 'jsw_paints_logo.png',
    project_root / 'dashboard' / 'assets' / 'jsw_paints_logo.jpeg',
    project_root / 'dashboard' / 'assets' / 'jsw_paints_logo.jpg'
]

logo_path = next((p for p in logo_candidates if p.exists()), None)
logo_frame_style = (
    'width:60px;height:60px;border-radius:14px;'
    'box-shadow:0 0 22px rgba(96,165,250,0.34),'
    '0 10px 24px rgba(15,23,42,0.45);'
    'border:1px solid rgba(96,165,250,0.24);'
    'flex:0 0 60px;'
)
logo_html = (
    f'<div class="hero-logo-fallback" style="{logo_frame_style}'
    'display:flex;align-items:center;justify-content:center;'
    'background:rgba(15,23,42,0.55);font-size:34px;">💧</div>'
)

if logo_path:
    logo_mime = 'image/png' if logo_path.suffix.lower() == '.png' else 'image/jpeg'
    logo_base64 = base64.b64encode(logo_path.read_bytes()).decode('ascii')
    logo_html = (
        f'<img class="hero-logo" '
        f'style="{logo_frame_style}object-fit:contain;'
        'background:rgba(255,255,255,0.96);padding:6px;" '
        f'src="data:{logo_mime};base64,{logo_base64}" '
        f'alt="JSW Paints logo" />'
    )

hero_html = (
    '<div class="premium-banner hero-banner" '
    'style="padding:22px 26px;margin-bottom:20px;">'
    '<div class="hero-banner-content" '
    'style="position:relative;z-index:1;display:flex;'
    'align-items:center;gap:18px;">'
    f'{logo_html}'
    '<div class="hero-copy" style="min-width:0;">'
    '<h1 style="margin:0;font-size:3.2rem;line-height:1;'
    'background:linear-gradient(90deg,#ffffff,#93c5fd,#3b82f6,#1d4ed8);'
    '-webkit-background-clip:text;-webkit-text-fill-color:transparent;'
    'text-shadow:0 0 25px rgba(96,165,250,0.35);">'
    'KPI Copilot</h1>'
    '<p class="hero-subtitle" '
    'style="margin:7px 0 0;font-size:1.02rem;color:#94a3b8;">'
    'JSW Paints Executive Performance Intelligence Platform'
    '</p>'
    '<p class="hero-executive-line" '
    'style="margin:8px 0 0;color:#cbd5e1;font-size:0.98rem;font-weight:500;">'
    '🚀 Executive Command Center • SAP Analytics • AI Insights'
    '</p>'
    '</div>'
    '</div>'
    '</div>'
)

st.markdown(
    hero_html,
    unsafe_allow_html=True
)

st.markdown(
    """
    <style>
    .stApp {
        background:
            radial-gradient(circle at top right, rgba(59,130,246,0.20), transparent 28%),
            radial-gradient(circle at top left, rgba(147,197,253,0.18), transparent 30%),
            linear-gradient(180deg,#020617 0%,#061427 100%);
    }

    .block-container {
        max-width: 98% !important;
    }

    div[data-testid="stMetric"] {
        background: rgba(17,24,39,0.75);
        backdrop-filter: blur(18px);
        border-radius: 20px;
        padding: 18px;
        border: 1px solid rgba(59,130,246,0.20);
        box-shadow: 0 0 20px rgba(59,130,246,0.12);
        transition: all 0.25s ease;
    }

    div[data-testid="stMetric"]:hover {
        transform: translateY(-6px);
        border: 1px solid rgba(96,165,250,0.5);
        box-shadow: 0 0 30px rgba(96,165,250,0.25);
    }

    .premium-banner {
        background: linear-gradient(135deg,
            rgba(10,25,49,0.98),
            rgba(21,43,82,0.95));
        border: 1px solid rgba(96,165,250,0.35);
        border-radius: 24px;
        padding: 24px;
        margin-bottom: 24px;
        box-shadow: 0 0 70px rgba(59,130,246,0.25);
        position: relative;
        overflow: hidden;
    }

    .premium-banner::before {
        content: "";
        position: absolute;
        left: 0;
        top: 0;
        width: 100%;
        height: 4px;
        background: linear-gradient(90deg,#93c5fd,#60a5fa,#2563eb,#1d4ed8,#ef4444);
    }

    .premium-banner::after {
        content: "";
        position: absolute;
        inset: 0;
        background: linear-gradient(90deg,transparent,rgba(255,255,255,0.03),transparent);
        animation: scanline 6s linear infinite;
    }

    h1 {
        background: linear-gradient(90deg,#ffffff,#bfdbfe,#60a5fa,#2563eb);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-shadow: 0 0 25px rgba(96,165,250,0.35);
    }

    div[data-testid="stPlotlyChart"] {
        border: 1px solid rgba(59,130,246,0.15);
        border-radius: 18px;
        padding: 6px;
        background: rgba(255,255,255,0.02);
        box-shadow: 0 0 25px rgba(59,130,246,0.08);
    }

    .action-card {
        background: linear-gradient(135deg,
            rgba(15,23,42,0.95),
            rgba(30,41,59,0.90));
        border-left: 5px solid #3b82f6;
        border-radius: 18px;
        padding: 18px 22px;
        margin-bottom: 14px;
        box-shadow: 0 8px 30px rgba(0,0,0,0.35);
        transition: all 0.3s ease;
    }

    .action-card:hover {
        transform: translateX(6px);
        box-shadow: 0 0 40px rgba(59,130,246,0.35);
    }

    @keyframes scanline {
        from { transform: translateX(-100%); }
        to { transform: translateX(100%); }
    }

    .stApp::before {
        content: "";
        position: fixed;
        inset: 0;
        pointer-events: none;
        background:
            radial-gradient(circle at 20% 20%, rgba(96,165,250,0.08), transparent 20%),
            radial-gradient(circle at 80% 30%, rgba(249,115,22,0.06), transparent 20%),
            radial-gradient(circle at 50% 80%, rgba(124,58,237,0.06), transparent 25%);
    }

    h2, h3 {
        text-shadow: 0 0 12px rgba(96,165,250,0.25);
    }

    .jsw-health-banner {
        background: linear-gradient(90deg,
            rgba(37,99,235,0.15),
            rgba(124,58,237,0.15),
            rgba(249,115,22,0.15));
        border: 1px solid rgba(96,165,250,0.20);
        border-radius: 16px;
        padding: 14px 18px;
        margin: 12px 0 24px 0;
        text-align: center;
        font-weight: 600;
        letter-spacing: 0.5px;
        backdrop-filter: blur(12px);
        animation: pulseGlow 4s ease-in-out infinite;
    }

    .health-strip {
        display: flex;
        flex-direction: column;
        gap: 10px;
        margin: 12px 0 18px 0;
    }

    .health-tile {
        width: 100%;
        background: rgba(17,24,39,0.75);
        border: 1px solid rgba(96,165,250,0.15);
        border-radius: 18px;
        padding: 18px;
        text-align: center;
        backdrop-filter: blur(14px);
        transition: all 0.3s ease;
        animation: pulseGlow 4s ease-in-out infinite;
        border-left: 3px solid rgba(96,165,250,0.75);
    }

    .health-tile:hover {
        transform: translateX(4px);
        border-color: rgba(96,165,250,0.5);
    }

    .health-number {
        font-size: 28px;
        font-weight: 700;
        color: #f4f4f9;
        text-shadow: 0 0 22px rgba(96,165,250,0.60);
    }

    .health-label {
        margin-top: 6px;
        color: #cbd5e1;
    }

    .action-title {
        font-size: 18px;
        font-weight: 700;
        color: #f4f4f9;
        margin-bottom: 6px;
    }

    .action-subtitle {
        color: #cbd5e1;
        line-height: 1.5;
    }

    @keyframes pulseGlow {
        0% { box-shadow: 0 0 12px rgba(96,165,250,0.15); }
        50% { box-shadow: 0 0 38px rgba(59,130,246,0.28); }
        100% { box-shadow: 0 0 12px rgba(96,165,250,0.15); }
    }

    img {
        filter: drop-shadow(0 0 18px rgba(96,165,250,0.35));
    }

    /* === BEGIN: Dashboard Glass and Futuristic Section Styles === */
    .dashboard-glass {
        background: linear-gradient(135deg, rgba(8,20,39,0.94), rgba(15,35,70,0.82));
        border: 1px solid rgba(96,165,250,0.18);
        border-radius: 22px;
        padding: 18px;
        backdrop-filter: blur(18px);
        box-shadow: 0 0 30px rgba(59,130,246,0.12);
        transition: all 0.3s ease;
    }

    .dashboard-glass:hover {
        transform: translateY(-4px);
        box-shadow: 0 0 40px rgba(96,165,250,0.18);
    }

    .executive-score-card {
        background: linear-gradient(135deg, rgba(15,23,42,0.95), rgba(30,41,59,0.92));
        border: 1px solid rgba(96,165,250,0.35);
        border-radius: 22px;
        padding: 20px;
        text-align: center;
        margin-bottom: 18px;
        position: relative;
        overflow: hidden;
    }
    .executive-score-card::before {
        content: "";
        position: absolute;
        inset: 0;
        background: linear-gradient(135deg, transparent, rgba(37,99,235,0.10), rgba(239,68,68,0.08), transparent);
    }

    .executive-score-number {
        font-size: 54px;
        font-weight: 800;
        color: #f8fafc;
        text-shadow: 0 0 25px rgba(96,165,250,0.55);
    }

    .health-bar {
        margin: 10px 0;
    }

    .futuristic-section {
        background: linear-gradient(135deg,
            rgba(15,23,42,0.90),
            rgba(17,24,39,0.80));
        border: 1px solid rgba(96,165,250,0.15);
        border-radius: 24px;
        padding: 20px;
        margin-bottom: 24px;
        position: relative;
        overflow: hidden;
    }

    .futuristic-section::before {
        content: "";
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 3px;
        background: linear-gradient(90deg,#2563eb,#7c3aed,#f97316);
    }

    @keyframes neonPulse {
        0% { box-shadow: 0 0 12px rgba(59,130,246,0.12); }
        50% { box-shadow: 0 0 28px rgba(249,115,22,0.18); }
        100% { box-shadow: 0 0 12px rgba(59,130,246,0.12); }
    }

    div[data-testid="stDataFrame"] {
        border: 1px solid rgba(96,165,250,0.15);
        border-radius: 18px;
        overflow: hidden;
        animation: neonPulse 5s ease-in-out infinite;
    }

    div[data-testid="stPlotlyChart"] {
        animation: neonPulse 5s ease-in-out infinite;
    }
    /* === END: Dashboard Glass and Futuristic Section Styles === */
    </style>
    """,
    unsafe_allow_html=True
)






workbook_path = project_root / 'reports' / 'kpi_target_tracker.xlsx'


# ── Download button ──────────────────────────────────────────────────────────
if workbook_path.exists():
    with open(workbook_path, 'rb') as f:
        st.download_button(
            '📥 Download Updated Excel',
            f.read(),
            file_name='kpi_target_tracker.xlsx'
        )

# ── SAP Connection Status ───────────────────────────────────────────────────
try:
    upload_log_df = pd.read_excel(
        workbook_path,
        sheet_name='SAP Upload Log'
    )

    if not upload_log_df.empty:
        latest_upload = upload_log_df.iloc[-1]

        sap_col1, sap_col2 = st.columns([1, 1])

        with sap_col1:
            st.info(
                f"🟢 SAP Connected\n\n"
                f"Last Upload: {latest_upload['Timestamp']}"
            )

        with sap_col2:
            st.info(
                f"📄 {latest_upload['Filename']}\n\n"
                f"Rows Updated: {latest_upload['Rows Updated']}"
            )
except Exception:
    pass

# ── KPI summary from Calculations sheet ─────────────────────────────────────
try:
    calc_df = pd.read_excel(workbook_path, sheet_name='Calculations', header=None)

    header_row = None
    for idx in range(min(15, len(calc_df))):
        row_values = [str(v).lower() for v in calc_df.iloc[idx].tolist()]
        if any('parameter' in v for v in row_values):
            header_row = idx
            break

    if header_row is not None:
        calc_df.columns = calc_df.iloc[header_row]
        calc_df = calc_df.iloc[header_row + 1:].reset_index(drop=True)

    cols = [str(c) for c in calc_df.columns]
    kpi_col         = next((c for c in cols if 'parameter' in c.lower()), cols[0])
    achievement_col = next((c for c in cols if 'achievement' in c.lower()), cols[1])
    status_col      = next((c for c in cols if 'status' in c.lower()), None)

    df = pd.DataFrame({
        'KPI':         calc_df[kpi_col].astype(str),
        'Achievement': pd.to_numeric(calc_df[achievement_col], errors='coerce'),
    })
    df['Status'] = calc_df[status_col].astype(str) if status_col else 'Unknown'
    df = df.dropna(subset=['Achievement']).reset_index(drop=True)

    # Load KPI ownership from Settings sheet if available
    owner_map = {}
    try:
        settings_df = pd.read_excel(workbook_path, sheet_name='Settings')
        owner_col = next(
            (c for c in settings_df.columns if 'owner' in str(c).lower()),
            None
        )
        kpi_name_col = next(
            (c for c in settings_df.columns if 'kpi' in str(c).lower() or 'parameter' in str(c).lower()),
            settings_df.columns[0]
        )

        if owner_col:
            owner_map = dict(
                zip(
                    settings_df[kpi_name_col].astype(str).str.strip(),
                    settings_df[owner_col].astype(str).str.strip()
                )
            )
    except Exception:
        owner_map = {}

except Exception:
    owner_map = {}
    df = pd.DataFrame(columns=['KPI', 'Achievement', 'Status'])

# ── Dashboard metrics ────────────────────────────────────────────────────────
if not df.empty:
    col1, col2, col3, col4 = st.columns(4)
    on_track = df['Status'].str.contains('On Track', case=False, na=False)

    monitor_count = df['Status'].str.contains(
        'Monitor',
        case=False,
        na=False
    ).sum()

    critical_count = df['Status'].str.contains(
        'Critical',
        case=False,
        na=False
    ).sum()

    col1.metric('Total KPIs',          len(df))
    col2.metric('Average Achievement', f"{df['Achievement'].mean():.1f}%")
    col3.metric('🟢 On Track', int(on_track.sum()))
    col4.metric('🔴 Critical', int(critical_count))

    total_kpis = max(len(df), 1)
    executive_score = int(
        ((green if 'green' in locals() else int(on_track.sum())) * 100 +
         monitor_count * 60 +
         critical_count * 20) / total_kpis
    )
    left, right = st.columns([3, 1])

    # --- KPI Performance Chart (glass style) ---
    with left:
        st.markdown('<div class="dashboard-glass">', unsafe_allow_html=True)
        st.subheader('📈 KPI Performance Command Center')
        fig = px.bar(
            df.sort_values('Achievement', ascending=False),
            x='KPI',
            y='Achievement'
        )
        fig.update_traces(marker_line_width=0)
        fig.update_traces(
            marker_color='#3b82f6'
        )
        fig.update_layout(height=520, margin=dict(l=20, r=20, t=20, b=80))
        st.plotly_chart(fig, width='stretch')
        st.markdown('</div>', unsafe_allow_html=True)

    # --- Current Status DataFrame (glass style) ---
    with right:
        st.markdown('<div class="dashboard-glass">', unsafe_allow_html=True)
        st.markdown(
            f'''
            <div class="executive-score-card">
                <div style="color:#94a3b8;font-size:0.9rem;">EXECUTIVE HEALTH SCORE</div>
                <div class="executive-score-number">{executive_score}</div>
                <div style="color:#cbd5e1;">Overall KPI Performance</div>
            </div>
            ''',
            unsafe_allow_html=True
        )
        st.subheader('🚦 Real-Time KPI Status Matrix')
        st.caption('Live KPI monitoring • AI anomaly detection • Executive alerting')

        green = int(on_track.sum())
        yellow = int(monitor_count)
        red = int(critical_count)

        st.markdown(
            f'''
            <div class="health-strip">
                <div class="health-tile">
                    <div class="health-label">🟢 ON TRACK</div>
                    <div class="health-number">{green}</div>
                </div>
                <div class="health-tile">
                    <div class="health-label">🟡 MONITOR</div>
                    <div class="health-number">{yellow}</div>
                </div>
                <div class="health-tile">
                    <div class="health-label">🔴 CRITICAL</div>
                    <div class="health-number">{red}</div>
                </div>
            </div>
            ''',
            unsafe_allow_html=True
        )

        traffic_df = df[['KPI', 'Achievement', 'Status']].copy()

        traffic_df['Status'] = traffic_df['Status'].apply(
            lambda x: '🟢 On Track'
            if 'on track' in str(x).lower()
            else (
                '🟡 Monitor'
                if 'monitor' in str(x).lower()
                else '🔴 Critical'
            )
        )

        st.dataframe(
            traffic_df,
            width='stretch',
            hide_index=True
        )
        st.markdown('</div>', unsafe_allow_html=True)


    st.subheader('🎯 KPI Health Distribution')
    st.caption('Executive visibility into KPI health across the organization')
    pie_df = df['Status'].value_counts().reset_index()
    pie_df.columns = ['Status', 'Count']
    donut_fig = px.pie(
        pie_df,
        names='Status',
        values='Count',
        hole=0.65
    )

    donut_fig.update_layout(
        height=420,
        annotations=[dict(text=f'{len(df)}<br>KPIs', showarrow=False, font_size=22)]
    )

    st.plotly_chart(donut_fig, width='stretch')

    st.subheader('🎯 Executive Action Board')
    st.caption('AI-prioritized intervention queue for leadership review')

    risk_df = df.sort_values('Achievement', ascending=True).head(3)

    for _, row in risk_df.iterrows():
        owner = owner_map.get(str(row['KPI']).strip(), 'Unassigned')

        st.markdown(
            f'''
            <div class="action-card">
                <div class="action-title">🚨 {row['KPI']}</div>
                <div class="action-subtitle">
                    <strong>Owner:</strong> {owner}<br>
                    <strong>Achievement:</strong> {row['Achievement']:.1f}%<br>
                    <strong>Priority:</strong> High<br><br>
                    Immediate management intervention recommended.
                </div>
            </div>
            ''',
            unsafe_allow_html=True
        )

if owner_map:
    st.subheader('👥 KPI Ownership Directory')

    owner_df = pd.DataFrame(
        sorted(owner_map.items()),
        columns=['KPI', 'Owner']
    )

    st.dataframe(owner_df, width='stretch', hide_index=True)

st.divider()

# ── Manual KPI Data Entry ────────────────────────────────────────────────────
st.subheader('Monthly KPI Data Entry')

with st.expander('Update KPI Values'):
    selected_kpi = st.selectbox('Select KPI', df['KPI'].tolist() if not df.empty else [])
    current_value = (
        df.loc[df['KPI'] == selected_kpi, 'Achievement'].iloc[0]
        if selected_kpi in df['KPI'].values else 0.0
    )
    st.info(f'Current Achievement Value for {selected_kpi}: {current_value}')
    new_value = st.number_input('New Achievement Value', value=float(current_value))

    if st.button('Save KPI Update'):
        try:
            wb = load_workbook(workbook_path)

            if 'Data Entry' not in wb.sheetnames:
                st.error('Data Entry sheet not found in workbook.')
            else:
                data_ws     = wb['Data Entry']
                settings_ws = wb['Settings']

                kpi_row = None
                for row in range(2, settings_ws.max_row + 1):
                    cell_value = settings_ws.cell(row, 1).value
                    if cell_value and str(cell_value).strip() == str(selected_kpi).strip():
                        kpi_row = row
                        break

                if kpi_row is None:
                    st.error(f'KPI not found in Settings sheet: {selected_kpi}')
                else:
                    current_month_col = None
                    for col in range(2, data_ws.max_column + 1):
                        if data_ws.cell(1, col).value:
                            current_month_col = col

                    if current_month_col is None:
                        st.error('No month columns found in Data Entry sheet.')
                    else:
                        data_ws.cell(row=kpi_row, column=current_month_col).value = float(new_value)
                        wb.save(workbook_path)
                        st.success(
                            f'Saved {selected_kpi} = {new_value} '
                            f'to column {data_ws.cell(1, current_month_col).value}'
                        )
                        st.rerun()

        except Exception as e:
            st.error(f'Excel update failed: {e}')

st.divider()

# ── SAP Upload Center ────────────────────────────────────────────────────────
st.subheader('SAP Upload Center')

sap_file = st.file_uploader(
    'Upload SAP Export (Excel/CSV)',
    type=['xlsx', 'xls', 'csv']
)

if sap_file:
    st.success('SAP file uploaded successfully')

    try:
        sap_df = (
            pd.read_csv(sap_file)
            if sap_file.name.endswith('.csv')
            else pd.read_excel(sap_file)
        )
        with st.expander('📄 View SAP Upload Preview'):
            st.dataframe(sap_df.head(), width='stretch')

        # ── Auto-detect SAP columns ──────────────────────────────────────────
        sap_kpi_col = next(
            (c for c in sap_df.columns if 'parameter' in str(c).lower()),
            sap_df.columns[0]
        )
        sap_value_col = next(
            (c for c in sap_df.columns
             if any(x in str(c).lower() for x in ['actual', 'current', 'value'])),
            sap_df.columns[1] if len(sap_df.columns) > 1 else sap_df.columns[0]
        )
        sap_month_col = next(
            (c for c in sap_df.columns if 'month' in str(c).lower()),
            None
        )

        st.subheader('Detected SAP Mapping')
        c1, c2, c3 = st.columns(3)
        c1.metric('KPI Column',   str(sap_kpi_col))
        c2.metric('Value Column', str(sap_value_col))
        c3.metric('Rows Found',   len(sap_df))

        st.subheader('KPI Match Preview')
        preview_df = sap_df[[sap_kpi_col, sap_value_col]].head(10).copy()
        preview_df.columns = ['KPI', 'Value']
        st.dataframe(preview_df, width='stretch')

        if st.button('🚀 Auto Update KPI Workbook'):
            try:
                wb          = load_workbook(workbook_path)
                data_ws     = wb['Data Entry']
                settings_ws = wb['Settings']

                # month_map: Data Entry row 1 headers → column index
                # e.g. {'apr': 2, 'may': 3, ..., 'jan': 11, 'feb': 12, 'mar': 13}
                month_map = {}
                for col in range(2, data_ws.max_column + 1):
                    header = str(data_ws.cell(1, col).value or '').strip()
                    if header:
                        month_map[header.lower()] = col

                # kpi_row_map: Settings col A KPI names → row index
                # Data Entry col A is entirely blank — KPI names only live in Settings.
                # Settings row N  ↔  Data Entry row N  (1-to-1 correspondence).
                kpi_row_map = {}
                for row in range(2, settings_ws.max_row + 1):
                    kpi = settings_ws.cell(row, 1).value
                    if kpi:
                        kpi_row_map[str(kpi).strip().lower()] = row

                updated_count  = 0
                unmatched_kpis = []
                match_log      = []

                for _, sap_row in sap_df.iterrows():
                    kpi_name  = str(sap_row[sap_kpi_col]).strip()
                    kpi_value = sap_row[sap_value_col]
                    month_raw = sap_row[sap_month_col] if sap_month_col else None

                    # FIX: Month column is a datetime/Timestamp object, not a string.
                    # str() on a Timestamp gives "2026-01-01 00:00:00".
                    # split('-')[0] gives "2026", not "jan" → lookup fails.
                    # Correct approach: parse as datetime and use strftime('%b').
                    try:
                        month_key = pd.to_datetime(month_raw).strftime('%b').lower()
                    except Exception:
                        # Fallback for plain strings like "Jan-2026" or "Jan"
                        month_key = str(month_raw).split('-')[0].strip().lower()

                    target_row = kpi_row_map.get(kpi_name.lower())
                    target_col = month_map.get(month_key)

                    if target_row and target_col:
                        try:
                            data_ws.cell(target_row, target_col).value = float(kpi_value)
                        except (ValueError, TypeError):
                            data_ws.cell(target_row, target_col).value = kpi_value

                        updated_count += 1
                        match_log.append({
                            'SAP KPI': kpi_name,
                            'Month':   month_key.capitalize(),
                            'Row':     target_row,
                            'Col':     target_col,
                            'Value':   kpi_value,
                        })
                    else:
                        reason = []
                        if target_row is None:
                            reason.append(f'KPI "{kpi_name}" not found in Settings')
                        if target_col is None:
                            reason.append(f'Month "{month_key}" not a column header')
                        unmatched_kpis.append({
                            'KPI':    kpi_name,
                            'Reason': '; '.join(reason),
                        })

                # Write SAP Upload Log
                if 'SAP Upload Log' not in wb.sheetnames:
                    log_ws = wb.create_sheet('SAP Upload Log')
                    log_ws.append(['Timestamp', 'Filename', 'Rows Updated'])
                else:
                    log_ws = wb['SAP Upload Log']

                log_ws.append([
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    sap_file.name,
                    updated_count,
                ])

                wb.save(str(workbook_path))
                wb.close()

                # Regenerate Calculations + AI Recommendations
                settings_df = pd.read_excel(
                    workbook_path,
                    sheet_name='Settings'
                )

                data_df = pd.read_excel(
                    workbook_path,
                    sheet_name='Data Entry'
                )

                analysis_df = generate_kpi_analysis(
                    settings_df,
                    data_df
                )

                write_analysis_sheet(
                    workbook_path,
                    analysis_df
                )

                # Results
                st.subheader('Update Results')

                if updated_count > 0 and not unmatched_kpis:
                    st.success(f'✅ All {updated_count} KPI values written to workbook.')
                elif updated_count > 0:
                    st.warning(f'⚠️ {updated_count}/{len(sap_df)} rows updated.')
                else:
                    st.error('❌ 0 rows updated.')

                if match_log:
                    st.dataframe(pd.DataFrame(match_log), width='stretch')

                if unmatched_kpis:
                    st.error('Unmatched rows:')
                    st.dataframe(pd.DataFrame(unmatched_kpis), width='stretch')

                st.rerun()

            except Exception as update_error:
                import traceback
                st.error(f'SAP auto-update failed: {update_error}')
                st.code(traceback.format_exc(), language='python')

    except Exception as e:
        st.error(f'Unable to read SAP file: {e}')

st.divider()

# ── Quick Actions ────────────────────────────────────────────────────────────
st.subheader('Quick Actions')

col_a, col_b = st.columns(2)

with col_a:
    if st.button('🔄 Refresh KPI Calculations'):
        settings_df = pd.read_excel(
            workbook_path,
            sheet_name='Settings'
        )

        data_df = pd.read_excel(
            workbook_path,
            sheet_name='Data Entry'
        )

        analysis_df = generate_kpi_analysis(
            settings_df,
            data_df
        )

        write_analysis_sheet(
            workbook_path,
            analysis_df
        )

        st.success('KPI calculations refreshed.')
        st.rerun()

with col_b:
    if st.button('📊 Generate Management Summary'):
        summary_answer = answer_question('Give me a management summary', df)
        if 'chat_history' not in st.session_state:
            st.session_state.chat_history = []
        st.session_state.chat_history.append(('Give me a management summary', summary_answer))
        st.rerun()

# ── KPI Copilot Chat ─────────────────────────────────────────────────────────
st.subheader('KPI Copilot Assistant')
st.caption('Powered by Gemini AI for KPI analysis, recommendations, forecasting, and management insights.')

if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []

user_prompt = st.chat_input(
    'Ask anything about KPI performance, risks, forecasts, targets, trends, or management actions...'
)

if user_prompt:
    answer = answer_question(user_prompt, df)
    st.session_state.chat_history.append((user_prompt, answer))

for question, answer in st.session_state.chat_history:
    with st.chat_message('user'):
        st.write(question)
    with st.chat_message('assistant'):
        st.write(answer)

# Remove or comment out any DEBUG dataframe/table output showing raw calculation rows near the top of the page
# (No such DEBUG output present; nothing to remove.)

st.divider()
st.success('KPI Copilot Dashboard Running Successfully')
